import json
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


EXCEL_FILE_PATH = (
    r"C:\Working files - No CF Projects\Documents - Personal"
    r"\RailRoadTracks\RRT Business\RRT Predictor"
    r"\rrt-predictor\data\RRT_Predictor_Upload.xlsx"
)

OUTPUT_JSON_PATH = Path("race_database.json")
BACKUP_FOLDER = Path("database_backups")
BACKUP_FOLDER.mkdir(exist_ok=True)

PREFERRED_SHEET_NAMES = [
    "Race Import Data",
    "Sheet1",
    "Export",
    "Races",
    "Data",
]


COLUMN_ALIASES = {
    "country": ["country", "nation"],
    "race_type": ["race type", "race_type", "code", "racing code"],
    "meeting_track": ["meeting / track", "meeting", "track", "venue", "racecourse"],
    "race_number": ["race number", "race no", "race", "race_number"],
    "race_date": ["race date", "date", "meeting date"],
    "race_time": ["race time", "time", "start time"],
    "race_name": ["race name", "event name"],
    "race_class": ["race class", "class"],
    "distance_m": ["distance m", "distance", "race distance"],
    "surface": ["surface"],
    "track_condition": ["track condition", "going", "track"],
    "weather": ["weather"],
    "horse_number": ["horse number", "runner number", "number", "no"],
    "horse_name": ["horse name", "runner", "runner name", "horse"],
    "barrier": ["barrier", "draw"],
    "weight_kg": ["weight kg", "weight", "handicap weight"],
    "jockey": ["jockey", "rider"],
    "jockey_win_percent": ["jockey win %", "jockey win percent", "jockey win"],
    "jockey_place_percent": ["jockey place %", "jockey place percent", "jockey place"],
    "trainer": ["trainer"],
    "trainer_win_percent": ["trainer win %", "trainer win percent", "trainer win"],
    "trainer_place_percent": ["trainer place %", "trainer place percent", "trainer place"],
    "recent_form": ["recent form", "form", "last 5", "last5", "last five"],
    "last_start_finish": ["last start finish", "last start", "last result"],
    "avg_prize_money": ["avg prize money", "average prize money", "prize money"],
    "track_condition_suitability": [
        "track condition suitability",
        "track suitability",
        "going suitability",
    ],
    "distance_suitability": ["distance suitability", "distance fit"],
    "market_rank": ["market rank", "rank", "odds rank"],
    "rrt_base_score": ["rrt base score", "base score", "rating"],
    "rrt_confidence_percent": ["rrt confidence %", "confidence", "confidence percent"],
    "notes": ["notes", "comment"],
    "meeting_id": ["meeting id", "meeting_id"],
    "upload_batch_id": ["upload batch id", "batch id"],
    "source_export_date": ["source export date", "export date"],
}


REQUIRED_COLUMNS = [
    "country",
    "race_type",
    "meeting_track",
    "race_number",
    "race_date",
    "race_time",
    "horse_name",
]


def clean_value(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, str):
        return value.strip()

    return value


def normalise_header(value: Any) -> str:
    return (
        str(value or "")
        .strip()
        .lower()
        .replace("\n", " ")
        .replace("\r", " ")
        .replace("\t", " ")
        .replace("_", " ")
        .replace("-", " ")
        .replace("/", " / ")
        .replace("%", "%")
    )


def normalise_key(value: Any) -> str:
    return (
        str(value or "")
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("/", "_")
        .replace("%", "percent")
        .replace("(", "")
        .replace(")", "")
        .replace("-", "_")
    )


def safe_float(value: Any, default: Optional[float] = None) -> Optional[float]:
    try:
        if value is None or value == "":
            return default

        cleaned = str(value).replace("%", "").replace("$", "").replace(",", "").strip()
        return float(cleaned)
    except Exception:
        return default


def safe_int(value: Any, default: Optional[int] = None) -> Optional[int]:
    try:
        if value is None or value == "":
            return default

        return int(float(str(value).replace(",", "").strip()))
    except Exception:
        return default


def choose_worksheet(workbook):
    for sheet_name in PREFERRED_SHEET_NAMES:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]

    return workbook[workbook.sheetnames[0]]


def build_header_map(raw_headers: List[Any]) -> Dict[str, int]:
    normalised_headers = [
        normalise_header(header)
        for header in raw_headers
    ]

    header_map = {}

    for canonical_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_normalised = normalise_header(alias)

            if alias_normalised in normalised_headers:
                header_map[canonical_name] = normalised_headers.index(alias_normalised)
                break

    return header_map


def validate_header_map(header_map: Dict[str, int]) -> List[str]:
    return [
        column
        for column in REQUIRED_COLUMNS
        if column not in header_map
    ]


def get_cell(row: tuple, header_map: Dict[str, int], key: str) -> Any:
    index = header_map.get(key)

    if index is None:
        return None

    if index >= len(row):
        return None

    return clean_value(row[index])


def derive_market_rank(row_data: Dict[str, Any]) -> Any:
    existing = row_data.get("market_rank")

    if existing is not None:
        return existing

    return None


def derive_default_score(value: Any, default: float = 50.0) -> float:
    score = safe_float(value, None)

    if score is None:
        return default

    return max(0, min(score, 100))


def normalise_row(row: tuple, header_map: Dict[str, int], upload_batch_id: str) -> Dict[str, Any]:
    row_data = {
        key: get_cell(row, header_map, key)
        for key in COLUMN_ALIASES.keys()
    }

    row_data["market_rank"] = derive_market_rank(row_data)
    row_data["upload_batch_id"] = row_data.get("upload_batch_id") or upload_batch_id
    row_data["source_export_date"] = (
        row_data.get("source_export_date")
        or datetime.now().strftime("%Y-%m-%d")
    )

    row_data["track_condition_suitability"] = derive_default_score(
        row_data.get("track_condition_suitability"),
        50,
    )

    row_data["distance_suitability"] = derive_default_score(
        row_data.get("distance_suitability"),
        50,
    )

    row_data["rrt_base_score"] = derive_default_score(
        row_data.get("rrt_base_score"),
        50,
    )

    row_data["jockey_win_percent"] = safe_float(row_data.get("jockey_win_percent"), 0)
    row_data["jockey_place_percent"] = safe_float(row_data.get("jockey_place_percent"), 0)
    row_data["trainer_win_percent"] = safe_float(row_data.get("trainer_win_percent"), 0)
    row_data["trainer_place_percent"] = safe_float(row_data.get("trainer_place_percent"), 0)
    row_data["last_start_finish"] = safe_int(row_data.get("last_start_finish"), 99)
    row_data["avg_prize_money"] = safe_float(row_data.get("avg_prize_money"), 0)
    row_data["market_rank"] = safe_int(row_data.get("market_rank"), 99)
    row_data["race_number"] = safe_int(row_data.get("race_number"), 0)
    row_data["horse_number"] = safe_int(row_data.get("horse_number"), None)
    row_data["barrier"] = safe_int(row_data.get("barrier"), None)
    row_data["weight_kg"] = safe_float(row_data.get("weight_kg"), None)

    return row_data


def backup_existing_database() -> Optional[Path]:
    if not OUTPUT_JSON_PATH.exists():
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_FOLDER / f"race_database_backup_{timestamp}.json"

    shutil.copy2(OUTPUT_JSON_PATH, backup_path)

    return backup_path


def load_excel_rows(file_path: str) -> Dict[str, Any]:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    workbook = load_workbook(file_path, data_only=True)
    worksheet = choose_worksheet(workbook)

    raw_headers = [cell.value for cell in worksheet[1]]
    header_map = build_header_map(raw_headers)
    missing_columns = validate_header_map(header_map)

    if missing_columns:
        return {
            "success": False,
            "message": "Excel file is missing required columns.",
            "sheet_name": worksheet.title,
            "missing_columns": missing_columns,
            "detected_columns": [str(header or "") for header in raw_headers],
            "rows": [],
        }

    upload_batch_id = f"upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    rows = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        item = normalise_row(row, header_map, upload_batch_id)

        if item.get("horse_name"):
            rows.append(item)

    return {
        "success": True,
        "message": "Excel rows loaded successfully.",
        "sheet_name": worksheet.title,
        "row_count": len(rows),
        "upload_batch_id": upload_batch_id,
        "rows": rows,
        "detected_columns": [str(header or "") for header in raw_headers],
    }


def build_database(rows: List[Dict[str, Any]], source_file: str, upload_batch_id: str) -> Dict[str, Any]:
    meetings = {}

    for row in rows:
        meeting_id = row.get("meeting_id")

        if not meeting_id:
            meeting_id = (
                f"{row.get('country')}_"
                f"{row.get('race_type')}_"
                f"{row.get('meeting_track')}_"
                f"{row.get('race_date')}"
            )

        meeting_id = normalise_key(meeting_id)

        race_number = row.get("race_number")
        race_id = f"{meeting_id}_race_{race_number}"

        if meeting_id not in meetings:
            meetings[meeting_id] = {
                "meeting_id": meeting_id,
                "country": row.get("country"),
                "race_type": row.get("race_type"),
                "meeting_track": row.get("meeting_track"),
                "race_date": str(row.get("race_date")),
                "source_export_date": str(row.get("source_export_date")),
                "upload_batch_id": upload_batch_id,
                "races": {},
            }

        if race_id not in meetings[meeting_id]["races"]:
            meetings[meeting_id]["races"][race_id] = {
                "race_id": race_id,
                "race_number": race_number,
                "race_time": str(row.get("race_time")),
                "race_name": row.get("race_name"),
                "race_class": row.get("race_class"),
                "distance_m": row.get("distance_m"),
                "surface": row.get("surface"),
                "track_condition": row.get("track_condition"),
                "weather": row.get("weather"),
                "runners": [],
            }

        runner = {
            "horse_number": row.get("horse_number"),
            "horse_name": row.get("horse_name"),
            "barrier": row.get("barrier"),
            "weight_kg": row.get("weight_kg"),
            "jockey": row.get("jockey"),
            "jockey_win_percent": row.get("jockey_win_percent"),
            "jockey_place_percent": row.get("jockey_place_percent"),
            "trainer": row.get("trainer"),
            "trainer_win_percent": row.get("trainer_win_percent"),
            "trainer_place_percent": row.get("trainer_place_percent"),
            "recent_form": row.get("recent_form"),
            "last_start_finish": row.get("last_start_finish"),
            "avg_prize_money": row.get("avg_prize_money"),
            "track_condition_suitability": row.get("track_condition_suitability"),
            "distance_suitability": row.get("distance_suitability"),
            "market_rank": row.get("market_rank"),
            "rrt_base_score": row.get("rrt_base_score"),
            "rrt_confidence_percent": row.get("rrt_confidence_percent"),
            "notes": row.get("notes"),
        }

        meetings[meeting_id]["races"][race_id]["runners"].append(runner)

    final_meetings = []

    for meeting in meetings.values():
        meeting["races"] = list(meeting["races"].values())
        meeting["race_count"] = len(meeting["races"])
        meeting["runner_count"] = sum(
            len(race["runners"])
            for race in meeting["races"]
        )
        final_meetings.append(meeting)

    return {
        "success": True,
        "generated_at": datetime.now().isoformat(),
        "source_file": source_file,
        "upload_batch_id": upload_batch_id,
        "meeting_count": len(final_meetings),
        "meetings": final_meetings,
    }


def import_race_data() -> Dict[str, Any]:
    loaded = load_excel_rows(EXCEL_FILE_PATH)

    if not loaded.get("success"):
        return loaded

    rows = loaded.get("rows", [])
    upload_batch_id = loaded.get("upload_batch_id")

    backup_path = backup_existing_database()

    database = build_database(
        rows=rows,
        source_file=EXCEL_FILE_PATH,
        upload_batch_id=upload_batch_id,
    )

    try:
        with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as file:
            json.dump(database, file, indent=2, ensure_ascii=False)

        database["backup_created"] = str(backup_path.resolve()) if backup_path else None
        database["sheet_name"] = loaded.get("sheet_name")
        database["row_count"] = loaded.get("row_count")
        database["detected_columns"] = loaded.get("detected_columns")

        return database

    except Exception as error:
        if backup_path and backup_path.exists():
            shutil.copy2(backup_path, OUTPUT_JSON_PATH)

        return {
            "success": False,
            "message": "Failed to write race database. Previous database restored if backup existed.",
            "error": str(error),
            "backup_restored": bool(backup_path),
        }


if __name__ == "__main__":
    result = import_race_data()

    print("Success:", result.get("success"))
    print("Message:", result.get("message", "Race database import completed."))
    print("Meetings:", result.get("meeting_count"))
    print("Rows:", result.get("row_count"))
    print("Output:", OUTPUT_JSON_PATH.resolve())

    if result.get("missing_columns"):
        print("Missing columns:", result.get("missing_columns"))